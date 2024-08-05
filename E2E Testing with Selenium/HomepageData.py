import openpyxl


class HomepageData:
#    test_homepageData = [{"FName":"TomCat", "Email":"abc@email.com", "Pwd":"abc123", "Gender":"Male"}, {"FName":"JerryMouse", "Email":"def@email.com", "Pwd":"def456", "Gender":"Female"}]

    @staticmethod
    def get_homepageData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\jcraj\\OneDrive\\pyDemo.xlsx")

        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]






